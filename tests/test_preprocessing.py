"""
Unit tests for src/preprocessing.py

Tests:
  - load_train_data: shape, types, no NaN
  - load_test_data:  correct row count, types, NaN positions, no crash on empty rows
  - SwaptionPreprocessor: fit/transform/inverse, winsorize bounds, numerical stability
  - parse_tenor_maturity / get_unique_tenors_maturities
"""

import os
import sys
import math
import tempfile
import numpy as np
import pytest
from openpyxl import Workbook

# ── Ensure project root is on sys.path ──
ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, ROOT)

from src.preprocessing import (
    load_train_data,
    load_test_data,
    SwaptionPreprocessor,
    parse_tenor_maturity,
    get_unique_tenors_maturities,
)


# ═══════════════════════════════════════════════════════════════
# Helpers: create minimal Excel files for isolated testing
# ═══════════════════════════════════════════════════════════════

def _make_train_xlsx(path, n_rows=10, n_cols=4):
    """Create a minimal train.xlsx with known values."""
    wb = Workbook()
    ws = wb.active
    headers = ["Date"] + [f"Tenor : {i+1}; Maturity : {j+1}"
                          for i in range(2) for j in range(n_cols // 2)]
    ws.append(headers)
    for r in range(n_rows):
        row = [f"2025-01-{r+1:02d}"] + [float(r * n_cols + c + 1) for c in range(n_cols)]
        ws.append(row)
    wb.save(path)


def _make_test_xlsx(path, n_price_cols=4, include_empty_rows=True):
    """Create a minimal test_template.xlsx mimicking the real format."""
    wb = Workbook()
    ws = wb.active
    price_headers = [f"Tenor : {i+1}; Maturity : {j+1}"
                     for i in range(2) for j in range(n_price_cols // 2)]
    headers = ["Type"] + price_headers + ["Date"]
    ws.append(headers)

    # 2 future-prediction rows (all NA)
    for d in ("2052-01-01", "2052-01-02"):
        ws.append(["Future prediction"] + ["NA"] * n_price_cols + [d])

    # 1 missing-data row (some NA, some numeric)
    vals = [1.5, "NA", 3.0, "NA"]
    ws.append(["Missing data"] + vals + ["2051-06-15"])

    # Optionally add empty rows (as real file does)
    if include_empty_rows:
        for _ in range(5):
            ws.append([])   # empty row

    wb.save(path)


# ═══════════════════════════════════════════════════════════════
# Tests: load_train_data
# ═══════════════════════════════════════════════════════════════

class TestLoadTrainData:
    """Tests for load_train_data()."""

    def test_shape_and_types(self, tmp_path):
        p = str(tmp_path / "train.xlsx")
        _make_train_xlsx(p, n_rows=10, n_cols=4)
        dates, cols, prices = load_train_data(p)

        assert len(dates) == 10
        assert len(cols) == 4
        assert prices.shape == (10, 4)
        assert prices.dtype == np.float32

    def test_no_nan_in_clean_data(self, tmp_path):
        p = str(tmp_path / "train.xlsx")
        _make_train_xlsx(p, n_rows=5, n_cols=4)
        _, _, prices = load_train_data(p)
        assert not np.isnan(prices).any()

    def test_handles_none_cells(self, tmp_path):
        """If a cell is None, load_train_data should still return (with NaN warning)."""
        p = str(tmp_path / "train.xlsx")
        wb = Workbook()
        ws = wb.active
        ws.append(["Date", "Col1", "Col2"])
        ws.append(["2025-01-01", 1.0, 2.0])
        ws.append(["2025-01-02", None, 4.0])  # one None cell
        wb.save(p)

        with pytest.warns(UserWarning, match="NaN"):
            dates, cols, prices = load_train_data(p)

        assert prices.shape == (2, 2)
        assert np.isnan(prices[1, 0])
        assert prices[1, 1] == 4.0

    def test_stops_at_empty_row(self, tmp_path):
        """Empty trailing rows should not produce extra data rows."""
        p = str(tmp_path / "train.xlsx")
        wb = Workbook()
        ws = wb.active
        ws.append(["Date", "C1"])
        ws.append(["2025-01-01", 1.0])
        ws.append([])                    # empty
        ws.append(["2025-01-03", 3.0])   # should be ignored
        wb.save(p)

        dates, _, prices = load_train_data(p)
        assert len(dates) == 1
        assert prices.shape == (1, 1)

    def test_real_file(self):
        """Smoke test against the actual DATASETS/train.xlsx if available."""
        real_path = os.path.join(ROOT, "DATASETS", "train.xlsx")
        if not os.path.exists(real_path):
            pytest.skip("Real train.xlsx not found")
        dates, cols, prices = load_train_data(real_path)
        assert prices.shape[0] > 0
        assert prices.shape[1] == 224
        assert not np.isnan(prices).any(), "Real training data should have no NaN"
        assert len(dates) == prices.shape[0]
        assert len(cols) == 224


# ═══════════════════════════════════════════════════════════════
# Tests: load_test_data
# ═══════════════════════════════════════════════════════════════

class TestLoadTestData:
    """Tests for load_test_data()."""

    def test_correct_row_count(self, tmp_path):
        p = str(tmp_path / "test.xlsx")
        _make_test_xlsx(p, n_price_cols=4, include_empty_rows=True)
        info, cols = load_test_data(p)
        # 2 future + 1 missing = 3 real rows; empty rows must not appear
        assert len(info) == 3

    def test_without_empty_rows(self, tmp_path):
        p = str(tmp_path / "test.xlsx")
        _make_test_xlsx(p, n_price_cols=4, include_empty_rows=False)
        info, cols = load_test_data(p)
        assert len(info) == 3

    def test_future_rows_all_nan(self, tmp_path):
        p = str(tmp_path / "test.xlsx")
        _make_test_xlsx(p, n_price_cols=4, include_empty_rows=True)
        info, _ = load_test_data(p)
        for row in info:
            if row["type"] == "Future prediction":
                assert np.all(np.isnan(row["values"]))

    def test_missing_data_nan_positions(self, tmp_path):
        p = str(tmp_path / "test.xlsx")
        _make_test_xlsx(p, n_price_cols=4)
        info, _ = load_test_data(p)
        missing_rows = [r for r in info if r["type"] == "Missing data"]
        assert len(missing_rows) == 1
        vals = missing_rows[0]["values"]
        # vals should be [1.5, NaN, 3.0, NaN]
        assert vals[0] == pytest.approx(1.5)
        assert np.isnan(vals[1])
        assert vals[2] == pytest.approx(3.0)
        assert np.isnan(vals[3])

    def test_column_count(self, tmp_path):
        p = str(tmp_path / "test.xlsx")
        _make_test_xlsx(p, n_price_cols=4)
        info, cols = load_test_data(p)
        assert len(cols) == 4
        for row in info:
            assert len(row["values"]) == 4

    def test_real_file(self):
        """Smoke test against the actual DATASETS/test_template.xlsx."""
        real_path = os.path.join(ROOT, "DATASETS", "test_template.xlsx")
        if not os.path.exists(real_path):
            pytest.skip("Real test_template.xlsx not found")
        info, cols = load_test_data(real_path)
        assert len(cols) == 224
        assert len(info) == 8, f"Expected 8 data rows, got {len(info)}"
        n_future  = sum(1 for r in info if r["type"] == "Future prediction")
        n_missing = sum(1 for r in info if r["type"] == "Missing data")
        assert n_future == 6
        assert n_missing == 2
        for row in info:
            assert len(row["values"]) == 224
            if row["type"] == "Future prediction":
                assert np.all(np.isnan(row["values"]))


# ═══════════════════════════════════════════════════════════════
# Tests: SwaptionPreprocessor
# ═══════════════════════════════════════════════════════════════

class TestSwaptionPreprocessor:
    """Tests for the preprocessing pipeline."""

    @pytest.fixture
    def data(self):
        """Synthetic 50×4 price matrix."""
        rng = np.random.RandomState(42)
        return rng.rand(50, 4).astype(np.float32) * 0.3 + 0.1  # values in [0.1, 0.4]

    def test_output_range(self, data):
        pp = SwaptionPreprocessor(winsorize_limits=(0.01, 0.01))
        normed = pp.fit_transform(data)
        assert normed.min() >= -0.01, f"Min below 0: {normed.min()}"
        assert normed.max() <= 1.01, f"Max above 1: {normed.max()}"

    def test_output_shape(self, data):
        pp = SwaptionPreprocessor()
        normed = pp.fit_transform(data)
        assert normed.shape == data.shape
        assert normed.dtype == np.float32

    def test_inverse_transform_roundtrip(self, data):
        pp = SwaptionPreprocessor()
        normed = pp.fit_transform(data)
        recovered = pp.inverse_transform(normed)
        # Winsorization clips extremes, so recovered ≈ clipped(data), not data.
        # The clipped version is the ground truth for the roundtrip.
        clipped = np.clip(data, pp.clip_lower_, pp.clip_upper_)
        np.testing.assert_allclose(recovered, clipped, atol=1e-3)

    def test_transform_uses_fitted_bounds(self, data):
        pp = SwaptionPreprocessor(winsorize_limits=(0.05, 0.05))
        pp.fit_transform(data)
        assert pp.clip_lower_ is not None
        assert pp.clip_upper_ is not None

        # New data with an extreme outlier should be clipped
        new_data = data[:5].copy()
        new_data[0, 0] = 999.0
        transformed = pp.transform(new_data)
        # Should not be wildly large — clipped at training upper bound
        assert transformed.max() < 10.0

    def test_winsorize_clips_extremes(self, data):
        pp = SwaptionPreprocessor(winsorize_limits=(0.10, 0.10))
        pp.fit_transform(data)
        # Lower/upper bounds should exclude the extreme 10%
        for col in range(data.shape[1]):
            assert pp.clip_lower_[col] >= np.min(data[:, col])
            assert pp.clip_upper_[col] <= np.max(data[:, col])

    def test_is_fitted_flag(self, data):
        pp = SwaptionPreprocessor()
        assert not pp.is_fitted
        pp.fit_transform(data)
        assert pp.is_fitted

    def test_transform_before_fit_raises(self, data):
        pp = SwaptionPreprocessor()
        with pytest.raises(AssertionError):
            pp.transform(data)

    def test_inverse_before_fit_raises(self, data):
        pp = SwaptionPreprocessor()
        with pytest.raises(AssertionError):
            pp.inverse_transform(data)

    def test_constant_column(self):
        """Constant-value column should not cause division by zero."""
        data = np.ones((20, 3), dtype=np.float32)
        pp = SwaptionPreprocessor()
        normed = pp.fit_transform(data)
        assert not np.isnan(normed).any()
        assert not np.isinf(normed).any()


# ═══════════════════════════════════════════════════════════════
# Tests: parse_tenor_maturity & get_unique_tenors_maturities
# ═══════════════════════════════════════════════════════════════

class TestTenorMaturityParsing:

    def test_parse_single(self):
        cols = ["Tenor : 5; Maturity : 10"]
        pairs = parse_tenor_maturity(cols)
        assert pairs == [(5.0, 10.0)]

    def test_parse_multiple(self):
        cols = [
            "Tenor : 1; Maturity : 0.0833333333333333",
            "Tenor : 2; Maturity : 0.25",
        ]
        pairs = parse_tenor_maturity(cols)
        assert len(pairs) == 2
        assert pairs[0][0] == pytest.approx(1.0)
        assert pairs[1][1] == pytest.approx(0.25)

    def test_get_unique(self):
        cols = [
            "Tenor : 1; Maturity : 10",
            "Tenor : 2; Maturity : 10",
            "Tenor : 1; Maturity : 20",
            "Tenor : 2; Maturity : 20",
        ]
        tenors, mats = get_unique_tenors_maturities(cols)
        assert tenors == [1.0, 2.0]
        assert mats == [10.0, 20.0]


# ═══════════════════════════════════════════════════════════════
# Entry point
# ═══════════════════════════════════════════════════════════════

if __name__ == "__main__":
    pytest.main([__file__, "-v", "--tb=short"])
