"""
Unit tests for src/autoencoder.py, src/hybrid_model.py, src/quantum_reservoir.py,
and src/predict.py

Covers:
  - SparseDenosingAE:  forward, encode, decode, encode_partial, shapes
  - AELoss:            sparsity penalty
  - ClassicalHead:     forward pass, shape
  - HybridLoss:        combined loss
  - make_windows:      windowing shapes & values
  - EnsembleQORC:      output dim, fallback features
  - write_predictions:  output format matches sample convention
  - predict helpers:    build_context, impute_missing
"""

import os
import sys
import math
import tempfile
import numpy as np
import torch
import pytest
from openpyxl import load_workbook

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, ROOT)

from src.autoencoder import SparseDenosingAE, AELoss
from src.hybrid_model import ClassicalHead, HybridLoss, make_windows
from src.quantum_reservoir import (
    EnsembleQORC, QuantumFeatureNormalizer,
    fock_output_size, extract_quantum_features,
)


# ═══════════════════════════════════════════════════════════════
# SparseDenosingAE
# ═══════════════════════════════════════════════════════════════

class TestSparseDenosingAE:

    @pytest.fixture
    def model(self):
        return SparseDenosingAE(input_dim=16, hidden_dims=(8,), latent_dim=4)

    def test_forward_shape(self, model):
        x = torch.rand(5, 16)
        x_hat, z = model(x, mask_ratio=0.0)
        assert x_hat.shape == (5, 16)
        assert z.shape == (5, 4)

    def test_decoder_output_in_01(self, model):
        x = torch.rand(5, 16)
        x_hat, _ = model(x)
        assert x_hat.min() >= 0.0
        assert x_hat.max() <= 1.0

    def test_elu_bottleneck_allows_negatives(self, model):
        """ELU activation should allow small negative latent values."""
        torch.manual_seed(123)
        x = torch.rand(100, 16)
        model.eval()
        with torch.no_grad():
            z = model.encode(x)
        # With random weights, ELU should produce some negative values
        assert z.min() < 0.0, "ELU should allow negatives"

    def test_masking_training_mode(self, model):
        model.train()
        x = torch.rand(10, 16)
        x_hat1, _ = model(x, mask_ratio=0.5)
        x_hat2, _ = model(x, mask_ratio=0.5)
        # Two forward passes with random masking should differ
        # (probabilistic, but with 50% masking this is virtually certain)
        assert not torch.allclose(x_hat1, x_hat2)

    def test_no_masking_eval_mode(self, model):
        model.eval()
        x = torch.rand(10, 16)
        with torch.no_grad():
            x_hat1, _ = model(x, mask_ratio=0.5)
            x_hat2, _ = model(x, mask_ratio=0.5)
        # In eval mode, masking is disabled → deterministic
        assert torch.allclose(x_hat1, x_hat2)

    def test_encode_partial_handles_nan(self, model):
        model.eval()
        x = torch.rand(3, 16)
        x[0, 5] = float("nan")
        x[1, 10] = float("nan")
        with torch.no_grad():
            z = model.encode_partial(x)
        assert not torch.isnan(z).any(), "encode_partial must replace NaN before encoding"

    def test_encode_decode_roundtrip(self, model):
        """Encoding then decoding should produce valid output (not testing accuracy)."""
        x = torch.rand(3, 16)
        model.eval()
        with torch.no_grad():
            z = model.encode(x)
            x_hat = model.decode(z)
        assert x_hat.shape == x.shape
        assert not torch.isnan(x_hat).any()


class TestAELoss:

    def test_sparsity_increases_loss(self):
        criterion_sparse = AELoss(sparsity_lambda=1.0)
        criterion_none   = AELoss(sparsity_lambda=0.0)

        x_hat = torch.rand(4, 8)
        x     = torch.rand(4, 8)
        z     = torch.ones(4, 4) * 5.0  # high activation

        loss_sparse, _, _ = criterion_sparse(x_hat, x, z)
        loss_none, _, _   = criterion_none(x_hat, x, z)

        assert loss_sparse > loss_none


# ═══════════════════════════════════════════════════════════════
# ClassicalHead & HybridLoss
# ═══════════════════════════════════════════════════════════════

class TestClassicalHead:

    def test_forward_shape(self):
        head = ClassicalHead(
            quantum_dim=100, classical_dim=60, latent_dim=10,
            hidden_dims=(32,), dropout=0.0,
        )
        q = torch.randn(4, 100)
        c = torch.randn(4, 60)
        out = head(q, c)
        assert out.shape == (4, 10)

    def test_gradient_flow(self):
        head = ClassicalHead(
            quantum_dim=50, classical_dim=30, latent_dim=5,
            hidden_dims=(16,), dropout=0.0,
        )
        q = torch.randn(2, 50, requires_grad=True)
        c = torch.randn(2, 30, requires_grad=True)
        out = head(q, c)
        loss = out.sum()
        loss.backward()
        assert q.grad is not None
        assert c.grad is not None


class TestHybridLoss:

    def test_surface_weight(self):
        loss_fn = HybridLoss(surface_weight=0.5)
        z_pred = torch.randn(3, 5)
        z_true = torch.randn(3, 5)
        s_pred = torch.rand(3, 16)
        s_true = torch.rand(3, 16)

        total, lat, surf = loss_fn(z_pred, z_true, s_pred, s_true)
        expected = lat + 0.5 * surf
        assert torch.allclose(total, expected)

    def test_zero_weight_ignores_surface(self):
        loss_fn = HybridLoss(surface_weight=0.0)
        z_pred = torch.randn(3, 5)
        z_true = torch.randn(3, 5)
        s_pred = torch.rand(3, 16)
        s_true = torch.rand(3, 16)

        total, lat, _ = loss_fn(z_pred, z_true, s_pred, s_true)
        assert torch.allclose(total, lat)


# ═══════════════════════════════════════════════════════════════
# make_windows
# ═══════════════════════════════════════════════════════════════

class TestMakeWindows:

    def test_shapes(self):
        latent = np.random.randn(20, 5).astype(np.float32)
        X, y, idx = make_windows(latent, window_size=3)
        expected_classical_dim = 5 * (3 + 1)  # window + delta
        assert X.shape == (17, expected_classical_dim)
        assert y.shape == (17, 5)
        assert len(idx) == 17

    def test_target_is_next_step(self):
        latent = np.arange(30).reshape(6, 5).astype(np.float32)
        X, y, idx = make_windows(latent, window_size=2)
        # y[0] should be latent[2] (the step after the first window)
        np.testing.assert_array_equal(y[0], latent[2])

    def test_delta_component(self):
        latent = np.random.randn(10, 3).astype(np.float32)
        X, y, idx = make_windows(latent, window_size=4)
        # Last 3 elements of context = delta = latent[3] - latent[2]
        expected_delta = latent[3] - latent[2]
        np.testing.assert_allclose(X[0, -3:], expected_delta, atol=1e-6)

    def test_window_size_1(self):
        latent = np.random.randn(5, 2).astype(np.float32)
        X, y, idx = make_windows(latent, window_size=1)
        assert X.shape == (4, 2 * 2)  # window(1)*2 + delta(2) = 4
        assert y.shape == (4, 2)


# ═══════════════════════════════════════════════════════════════
# EnsembleQORC & QuantumFeatureNormalizer
# ═══════════════════════════════════════════════════════════════

class TestFockOutputSize:

    def test_known_values(self):
        # C(14, 3) = 364
        assert fock_output_size(12, 3, use_fock=True) == math.comb(14, 3)
        # C(13, 4) = 715
        assert fock_output_size(10, 4, use_fock=True) == math.comb(13, 4)
        # No-bunching: C(m, n)
        assert fock_output_size(12, 3, use_fock=False) == math.comb(12, 3)

    def test_total_1215(self):
        total = (fock_output_size(12, 3) +
                 fock_output_size(10, 4) +
                 fock_output_size(16, 2))
        assert total == 1215


class TestEnsembleQORC:

    def test_fallback_output_shape(self):
        configs = [
            {"n_modes": 4, "n_photons": 2, "seed": 1},
            {"n_modes": 3, "n_photons": 1, "seed": 2},
        ]
        ensemble = EnsembleQORC(input_dim=10, configs=configs, device="cpu")
        x = torch.randn(5, 10)
        out = ensemble(x)
        expected_dim = fock_output_size(4, 2) + fock_output_size(3, 1)
        assert out.shape == (5, expected_dim)

    def test_no_nan_inf(self):
        configs = [{"n_modes": 4, "n_photons": 2, "seed": 42}]
        ensemble = EnsembleQORC(input_dim=8, configs=configs, device="cpu")
        x = torch.randn(10, 8)
        out = ensemble(x)
        assert not torch.isnan(out).any()
        assert not torch.isinf(out).any()

    def test_deterministic_fallback(self):
        """Fallback features should be deterministic for same input."""
        configs = [{"n_modes": 4, "n_photons": 2, "seed": 42}]
        ensemble = EnsembleQORC(input_dim=8, configs=configs, device="cpu")
        x = torch.randn(3, 8)
        out1 = ensemble(x)
        out2 = ensemble(x)
        assert torch.allclose(out1, out2)


class TestQuantumFeatureNormalizer:

    def test_fit_transform_normalizes(self):
        data = np.random.randn(50, 10).astype(np.float32) * 3 + 5
        norm = QuantumFeatureNormalizer()
        normed = norm.fit_transform(data)
        assert abs(normed.mean()) < 0.1
        assert abs(normed.std() - 1.0) < 0.1

    def test_transform_uses_fitted_stats(self):
        data = np.random.randn(50, 10).astype(np.float32)
        norm = QuantumFeatureNormalizer()
        norm.fit_transform(data)

        new_data = np.ones((5, 10), dtype=np.float32) * 100
        transformed = norm.transform(new_data)
        # Should be far from 0 since 100 is far from training mean
        assert transformed.mean() > 5.0


# ═══════════════════════════════════════════════════════════════
# write_predictions
# ═══════════════════════════════════════════════════════════════

class TestWritePredictions:

    def test_output_format(self, tmp_path):
        from src.predict import write_predictions
        from datetime import datetime

        price_cols = ["Col1", "Col2", "Col3"]
        test_info = [
            {"type": "Future prediction", "date": datetime(2052, 1, 1), "values": None},
            {"type": "Missing data", "date": datetime(2051, 6, 15), "values": None},
        ]
        predictions = [
            np.array([0.1, 0.2, 0.3]),
            np.array([0.4, 0.5, 0.6]),
        ]
        out = str(tmp_path / "pred.xlsx")
        write_predictions(test_info, predictions, price_cols, out)

        # Verify output file
        wb = load_workbook(out)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))

        # Header: [Col1, Col2, Col3, Date, Type]
        assert rows[0] == ("Col1", "Col2", "Col3", "Date", "Type")

        # Row 1: Future → "Complete"
        assert rows[1][-1] == "Complete"
        assert rows[1][0] == pytest.approx(0.1)

        # Row 2: Missing → "Missing Data"
        assert rows[2][-1] == "Missing Data"

        wb.close()

    def test_all_values_numeric(self, tmp_path):
        from src.predict import write_predictions
        from datetime import datetime

        price_cols = ["A", "B"]
        test_info = [{"type": "Future prediction", "date": datetime(2052, 1, 1), "values": None}]
        predictions = [np.array([1.234, 5.678])]
        out = str(tmp_path / "pred.xlsx")
        write_predictions(test_info, predictions, price_cols, out)

        wb = load_workbook(out)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        # All price cells should be float
        for v in rows[1][:2]:
            assert isinstance(v, float)
        wb.close()


# ═══════════════════════════════════════════════════════════════
# predict helpers: build_context
# ═══════════════════════════════════════════════════════════════

class TestBuildContext:

    def test_shape(self):
        from src.predict import build_context
        window = np.random.randn(5, 10).astype(np.float32)
        ctx = build_context(window)
        expected_len = 5 * 10 + 10  # flat + delta
        assert ctx.shape == (expected_len,)

    def test_delta_values(self):
        from src.predict import build_context
        window = np.arange(20).reshape(4, 5).astype(np.float32)
        ctx = build_context(window)
        # delta = window[-1] - window[-2] = [15,16,17,18,19] - [10,11,12,13,14] = [5,5,5,5,5]
        np.testing.assert_array_equal(ctx[-5:], [5, 5, 5, 5, 5])


# ═══════════════════════════════════════════════════════════════
# predict helpers: impute_missing
# ═══════════════════════════════════════════════════════════════

class TestImputeMissing:

    def test_observed_values_preserved(self):
        from src.predict import impute_missing
        from src.autoencoder import SparseDenosingAE
        from src.preprocessing import SwaptionPreprocessor

        # Build minimal model and preprocessor
        model = SparseDenosingAE(input_dim=8, hidden_dims=(4,), latent_dim=2)
        model.eval()

        data = np.random.rand(10, 8).astype(np.float32) * 0.3 + 0.1
        pp = SwaptionPreprocessor()
        pp.fit_transform(data)

        partial = data[0].copy()
        partial[3] = np.nan
        partial[5] = np.nan

        full = impute_missing(partial, pp, model, torch.device("cpu"),
                              train_prices=data)

        # Observed values must be unchanged
        for i in range(8):
            if i not in (3, 5):
                assert full[i] == pytest.approx(partial[i], abs=1e-5)

        # NaN positions must be filled
        assert not np.isnan(full[3])
        assert not np.isnan(full[5])


# ═══════════════════════════════════════════════════════════════
# Entry point
# ═══════════════════════════════════════════════════════════════

if __name__ == "__main__":
    pytest.main([__file__, "-v", "--tb=short"])
