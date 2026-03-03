"""Quick inspection of test.xlsx vs train.xlsx"""
from openpyxl import load_workbook
import numpy as np

# Load test data
wb = load_workbook('DATASETS/test.xlsx', read_only=True)
ws = wb.active
headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
print(f'TEST  - Columns: {len(headers)}')
print(f'  First 5: {headers[:5]}')
print(f'  Last 5: {headers[-5:]}')

rows = []
for row in ws.iter_rows(min_row=2, values_only=True):
    row = list(row)
    if len(row) == 0 or row[0] is None:
        break
    rows.append(row)
wb.close()

print(f'  Rows: {len(rows)}')
for i, r in enumerate(rows):
    vals = [v for v in r[1:] if v is not None]
    print(f'    Row {i}: date={r[0]}, non-null={len(vals)}, sample={vals[:3]}')

# Load train data for comparison
wb2 = load_workbook('DATASETS/train.xlsx', read_only=True)
ws2 = wb2.active
headers2 = [cell.value for cell in next(ws2.iter_rows(min_row=1, max_row=1))]
train_rows = []
for row in ws2.iter_rows(min_row=2, values_only=True):
    row = list(row)
    if len(row) == 0 or row[0] is None:
        break
    train_rows.append(row)
wb2.close()

print(f'\nTRAIN - Rows: {len(train_rows)}, Columns: {len(headers2)}')
print(f'  Last date: {train_rows[-1][0]}')
print(f'  Headers match: {headers == headers2}')
print(f'  Last 3 train dates: {[r[0] for r in train_rows[-3:]]}')
