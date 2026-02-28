"""
Preprocessing pipeline for swaption data.
Handles: data loading, outlier treatment, normalization, inverse transforms.
"""

import numpy as np
from openpyxl import load_workbook


class SwaptionPreprocessor:
    """Robust preprocessing: winsorize → RobustScaler (manual) → MinMaxScaler (manual).

    Implements scalers manually (numpy only) so it runs anywhere without sklearn.
    """

    def __init__(self, winsorize_limits=(0.01, 0.01)):
        self.winsorize_limits = winsorize_limits
        self.is_fitted = False
        # RobustScaler params (per column)
        self.median_ = None
        self.iqr_ = None
        # MinMaxScaler params (per column)
        self.min_ = None
        self.range_ = None

    def _winsorize(self, data):
        """Clip each column to [lower_pct, upper_pct] percentiles."""
        result = np.copy(data)
        lo, hi = self.winsorize_limits
        for col in range(result.shape[1]):
            lower = np.percentile(result[:, col], lo * 100)
            upper = np.percentile(result[:, col], (1 - hi) * 100)
            result[:, col] = np.clip(result[:, col], lower, upper)
        return result

    def fit_transform(self, data):
        """
        Fit on training data and transform.
        Args:
            data: np.ndarray of shape (n_timesteps, 224)
        Returns:
            Normalized data in [0, 1], shape (n_timesteps, 224)
        """
        # Step 1: Winsorize outliers
        data_clean = self._winsorize(data)

        # Step 2: RobustScaler (median / IQR)
        self.median_ = np.median(data_clean, axis=0)
        q75 = np.percentile(data_clean, 75, axis=0)
        q25 = np.percentile(data_clean, 25, axis=0)
        self.iqr_ = q75 - q25
        self.iqr_[self.iqr_ == 0] = 1.0  # avoid division by zero
        data_robust = (data_clean - self.median_) / self.iqr_

        # Step 3: MinMax to [0, 1]
        self.min_ = data_robust.min(axis=0)
        self.range_ = data_robust.max(axis=0) - self.min_
        self.range_[self.range_ == 0] = 1.0
        data_normalized = (data_robust - self.min_) / self.range_

        self.is_fitted = True
        return data_normalized.astype(np.float32)

    def transform(self, data):
        """Transform using fitted scalers."""
        assert self.is_fitted, "Call fit_transform first"
        data_clean = self._winsorize(data)
        data_robust = (data_clean - self.median_) / self.iqr_
        data_normalized = (data_robust - self.min_) / self.range_
        return data_normalized.astype(np.float32)

    def inverse_transform(self, data):
        """Reverse normalization to recover original scale prices."""
        assert self.is_fitted, "Call fit_transform first"
        try:
            import torch
            if isinstance(data, torch.Tensor):
                data = data.detach().cpu().numpy()
        except ImportError:
            pass
        data = np.asarray(data, dtype=np.float64)
        # Reverse MinMax
        data_robust = data * self.range_ + self.min_
        # Reverse RobustScaler
        data_original = data_robust * self.iqr_ + self.median_
        return data_original


def load_train_data(path="DATASETS/train.xlsx"):
    """
    Load training data.
    Returns:
        dates: list of date strings
        price_columns: list of column header strings
        prices: np.ndarray of shape (494, 224)
    """
    wb = load_workbook(path, read_only=True)
    ws = wb.active

    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    price_columns = headers[1:]  # Skip 'Date'

    dates = []
    prices = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        dates.append(row[0])
        prices.append([float(v) for v in row[1:]])

    wb.close()
    return dates, price_columns, np.array(prices, dtype=np.float32)


def load_test_data(path="DATASETS/test_template.xlsx"):
    """
    Load test template.
    Returns:
        test_info: list of dicts with 'type', 'date', 'values' (with NaN for missing)
        price_columns: list of column header strings (224 cols)
    """
    wb = load_workbook(path, read_only=True)
    ws = wb.active

    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    # Format: Type, 224 price cols, Date
    price_columns = headers[1:-1]

    test_info = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        row_type = row[0]
        row_date = row[-1]
        values = []
        for v in row[1:-1]:
            if v == "NA" or v is None:
                values.append(np.nan)
            else:
                values.append(float(v))
        test_info.append({
            "type": row_type,
            "date": row_date,
            "values": np.array(values, dtype=np.float32),
        })

    wb.close()
    return test_info, price_columns


def parse_tenor_maturity(columns):
    """Parse 'Tenor : X; Maturity : Y' headers into (tenor, maturity) tuples."""
    pairs = []
    for col in columns:
        parts = col.split(";")
        tenor = float(parts[0].split(":")[1].strip())
        maturity = float(parts[1].split(":")[1].strip())
        pairs.append((tenor, maturity))
    return pairs


def get_unique_tenors_maturities(columns):
    """Get sorted unique tenor and maturity values from column headers."""
    pairs = parse_tenor_maturity(columns)
    tenors = sorted(set(t for t, m in pairs))
    maturities = sorted(set(m for t, m in pairs))
    return tenors, maturities


# ============================================================
# Smoke test
# ============================================================
if __name__ == "__main__":
    print("=" * 60)
    print("SMOKE TEST: Preprocessing Pipeline")
    print("=" * 60)

    # 1. Load data
    print("\n[1] Loading training data...")
    dates, columns, prices = load_train_data()
    print(f"  Dates: {len(dates)} ({dates[0]} to {dates[-1]})")
    print(f"  Price columns: {len(columns)}")
    print(f"  Prices shape: {prices.shape}")
    print(f"  Price range: [{prices.min():.6f}, {prices.max():.6f}]")

    # 2. Parse tenor/maturity grid
    print("\n[2] Parsing tenor/maturity grid...")
    tenors, maturities = get_unique_tenors_maturities(columns)
    print(f"  Tenors ({len(tenors)}): {tenors}")
    print(f"  Maturities ({len(maturities)}): {maturities}")
    print(f"  Grid: {len(tenors)} x {len(maturities)} = {len(tenors)*len(maturities)}")

    # 3. Preprocess
    print("\n[3] Preprocessing...")
    preprocessor = SwaptionPreprocessor()
    prices_norm = preprocessor.fit_transform(prices)
    print(f"  Normalized range: [{prices_norm.min():.6f}, {prices_norm.max():.6f}]")

    # 4. Verify invertibility
    print("\n[4] Verifying inverse transform...")
    prices_recovered = preprocessor.inverse_transform(prices_norm)
    max_error = np.max(np.abs(prices - prices_recovered))
    print(f"  Max reconstruction error: {max_error:.10f}")
    assert max_error < 0.01, f"Inverse transform error too large: {max_error}"
    print("  PASSED")

    # 5. Load test data
    print("\n[5] Loading test data...")
    test_info, test_columns = load_test_data()
    for i, info in enumerate(test_info):
        n_missing = np.sum(np.isnan(info["values"]))
        print(f"  Row {i+1}: {info['type']}, date={info['date']}, missing={n_missing}/224")

    print("\n" + "=" * 60)
    print("ALL SMOKE TESTS PASSED")
    print("=" * 60)
